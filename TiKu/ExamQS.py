import tkinter as tk
import pandas as pd
from docxtpl import DocxTemplate, RichText
from openpyxl import load_workbook
from docx import Document
from pathlib import Path  # 导入pathlib模块中的Path类
import os
import sys
import numpy as np

BASE_DIR = os.path.dirname(os.path.realpath(sys.argv[0]))
df = pd.read_excel(os.path.join(BASE_DIR, "TK.xlsx"), sheet_name='Sheet1')
dfDF = pd.DataFrame(df)

def GUI(a, b, i):
    c = tk.Checkbutton(root, text=a, variable=b, onvalue=1, offvalue=0)
    c.grid(column=0, row=i, ipadx=5, pady=5)
    l_t = tk.Label(root, text='每题分值：')
    l_t.grid(column=1, row=i, ipadx=5, pady=5)
    l_tt = tk.Label(root, text='题目数量：')
    l_tt.grid(column=3, row=i, ipadx=5, padx =20, pady=5)

def twj():
    global dfDF
    df = pd.read_excel(os.path.join(BASE_DIR, "TK.xlsx"), sheet_name='Sheet1')
    dfDF = pd.DataFrame(df)
    a = [v1.get(), v2.get(), v3.get(), v4.get(), v5.get(), v6.get(), v7.get()]
    b = [EntryList2[0].get(), EntryList2[1].get(), EntryList2[2].get(), EntryList2[3].get(), EntryList2[4].get(), EntryList2[5].get(), EntryList2[6].get()]
    c1 = []
    c2 = []
    r1 = []
    r2 = []
    for j in [1,2]:   #生成两份试卷
        for i in [1, 2, 3, 4, 5, 6, 7]:  #对各个题型的设置
            if a[i - 1] == 1:   #对选取的题型执行下面的操作
                test = 0
                while test < int(b[i - 1]):  #题量的控制
                    randomK = dfDF.sample()  #随机抽取一个试题
                    indicator = randomK["知识点标号"].values  #查询抽取的试题的知识点标号
                    screenK = dfDF.loc[dfDF["知识点标号"].values == indicator]  #将该知识点对应的所有试题提取出来
                    aa = screenK["题型"].values  #统计这些试题的各自对应的题型
                    x = any(_ == i for _ in aa)  #如果题型中含有所需的题型时，执行下面的操作
                    if x:
                        screenKK = screenK.loc[screenK["题型"].values == i]  #将所需的题型的试题提取出来，题目数量大于等于1
                        exam_a = screenKK.sample()  #随机抽取一个试题
                        indexname = exam_a.index  #读取这些提取出来的试题的索引
                        dfDF = dfDF.drop(indexname)
                        writer = pd.ExcelWriter(os.path.join(BASE_DIR, "test\\test" + str(j) + "\\test-" + str(i) + "-" + str(j) + ".xlsx"),engine="openpyxl", mode="a", if_sheet_exists="new")
                        exam_a.to_excel(writer, sheet_name='Sheet1', index=False, header=True)
                        writer.save()
                        test = test + 1
                    else:
                        test = test
                        pass

                iris = pd.read_excel(os.path.join(BASE_DIR, "test\\test" + str(j) + "\\test-" + str(i) + "-" + str(j) + ".xlsx"),None)  # 读入数据文件
                keys = list(iris.keys())
                iris_concat = pd.DataFrame()
                for u in keys[1:int(b[i - 1]) + 1]:
                    iris_a = iris[u]
                    iris_concat = pd.concat([iris_concat, iris_a], sort=False)
                    iris_concat = iris_concat.reset_index(drop=True)
                iris_concat.to_excel(os.path.join(BASE_DIR, "test\\test" + str(j) + "\\test-" + str(i) + "-" + str(j) + ".xlsx"),sheet_name='Sheet1')  # 数据保存路


def cal_add():
    text.delete("1.0", tk.END)
    a = [v1.get(), v2.get(), v3.get(), v4.get(), v5.get(), v6.get(), v7.get()]
    b = [EntryList2[0].get(), EntryList2[1].get(), EntryList2[2].get(), EntryList2[3].get(), EntryList2[4].get(), EntryList2[5].get(), EntryList2[6].get()]
    c = [EntryList1[0].get(), EntryList1[1].get(), EntryList1[2].get(), EntryList1[3].get(), EntryList1[4].get(), EntryList1[5].get(), EntryList1[6].get()]
    list1 = []
    for i in [1, 2, 3, 4, 5, 6, 7]:
        if a[i - 1] == 1:
            var_entry1 = c[i - 1]  # 每题得分
            var_entry2 = b[i - 1]  # 题量
            add = int(var_entry1) * int(var_entry2)
            list1.append(add)
        else:
            var_entry1 = 0
            var_entry2 = 0
    sum_list1 = sum(list1)
    text.insert('insert', sum_list1)


def word():
    dun = "."
    kongG = " "
    for y in [1, 2]:
        exec("dictt%s1 = {}" % y)
        exec("dictt%s2 = {}" % y)
        exec("dictt%s3 = {}" % y)
        exec("dictt%s4 = {}" % y)
        exec("dictt%s5 = {}" % y)
        exec("dictt%s6 = {}" % y)
        exec("dictt%s7 = {}" % y)
    # 生成dictt11和dictt21两个字典
    if v1.get() == 1:
        for x1 in [1, 2]:
            wb_1 = load_workbook("test\\test" + str(x1) + "\\test-1-" + str(x1) + ".xlsx")
            sheets = wb_1.sheetnames
            aaa = wb_1[sheets[0]]
            for i in range(int(EntryList2[0].get())):
                mediam1 = aaa.cell(row=i + 2, column=4).value
                mediam1 = str(i + 1) + dun + kongG + str(mediam1)
                exec("dictt%s1['testa' + str(i+1)] = mediam1" % x1)
    # 生成dictt12和dictt22两个字典
    if v2.get() == 1:
        for x2 in [1, 2]:
            wb_2 = load_workbook("test\\test" + str(x2) + "\\test-2-" + str(x2) + ".xlsx")
            sheets2 = wb_2.sheetnames
            bbb = wb_2[sheets2[0]]

            for i in range(int(EntryList2[1].get())):
                mediam2 = bbb.cell(row=i + 2, column=4).value
                mediam2 = str(i + 1) + dun + kongG + str(mediam2)
                exec("dictt%s2['testb' + str(i+1)] = mediam2" % x2)
    # 生成dictt13和dictt23两个字典
    if v3.get() == 1:
        for x3 in [1, 2]:
            wb_3 = load_workbook("test\\test" + str(x3) + "\\test-3-" + str(x3) + ".xlsx")
            sheets3 = wb_3.sheetnames
            ccc = wb_3[sheets3[0]]

            for i in range(int(EntryList2[2].get())):
                mediam3 = ccc.cell(row=i + 2, column=4).value
                mediam3 = str(i + 1) + dun + kongG + str(mediam3)
                exec("dictt%s3['testc' + str(i+1)] = mediam3" % x3)
    # 生成dictt14和dictt24两个字典
    if v4.get() == 1:
        for x4 in [1, 2]:
            wb_4 = load_workbook("test\\test" + str(x4) + "\\test-4-" + str(x4) + ".xlsx")
            sheets4 = wb_4.sheetnames
            ddd = wb_4[sheets4[0]]

            for i in range(int(EntryList2[3].get())):
                mediam4 = ddd.cell(row=i + 2, column=4).value
                mediam4 = str(i + 1) + dun + kongG + str(mediam4)
                exec("dictt%s4['testd' + str(i+1)] = mediam4" % x4)
    # 生成dictt15和dictt25两个字典
    if v5.get() == 1:
        for x5 in [1, 2]:
            wb_5 = load_workbook("test\\test" + str(x5) + "\\test-5-" + str(x5) + ".xlsx")
            sheets5 = wb_5.sheetnames
            eee = wb_5[sheets5[0]]

            for i in range(int(EntryList2[4].get())):
                mediam5 = eee.cell(row=i + 2, column=4).value
                mediam5 = str(i + 1) + dun + kongG + str(mediam5)
                exec("dictt%s5['teste' + str(i+1)] = mediam5" % x5)
    # 生成dictt16和dictt26两个字典
    if v6.get() == 1:
        for x6 in [1, 2]:
            wb_6 = load_workbook("test\\test" + str(x6) + "\\test-6-" + str(x6) + ".xlsx")
            sheets6 = wb_6.sheetnames
            fff = wb_6[sheets6[0]]

            for i in range(int(EntryList2[5].get())):
                mediam6 = fff.cell(row=i + 2, column=4).value
                mediam6 = str(i + 1) + dun + kongG + str(mediam6)
                exec("dictt%s6['testf' + str(i+1)] = mediam6" % x6)
    # 生成dictt16和dictt26两个字典
    if v7.get() == 1:
        for x7 in [1, 2]:
            wb_7 = load_workbook("test\\test" + str(x7) + "\\test-7-" + str(x7) + ".xlsx")
            sheets7 = wb_7.sheetnames
            ggg = wb_7[sheets7[0]]

            for i in range(int(EntryList2[6].get())):
                mediam7 = ggg.cell(row=i + 2, column=4).value
                mediam7 = str(i + 1) + dun + kongG + str(mediam7)
                exec("dictt%s7['testg' + str(i+1)] = mediam7" % x7)

    doc1 = DocxTemplate("试卷模板.docx")
    doc2 = DocxTemplate("试卷模板.docx")
    for i in [1, 2]:
        exec("dictt%s6.update(dictt%s7)" % (i, i))
        exec("A%s=dictt%s6" % (i, i))
        exec("dictt%s5.update(A%s)" % (i, i))
        exec("B%s=dictt%s5" % (i, i))
        exec("dictt%s4.update(B%s)" % (i, i))
        exec("C%s=dictt%s4" % (i, i))
        exec("dictt%s3.update(C%s)" % (i, i))
        exec("D%s=dictt%s3" % (i, i))
        exec("dictt%s2.update(D%s)" % (i, i))
        exec("E%s=dictt%s2" % (i, i))
        exec("dictt%s1.update(E%s)" % (i, i))
        exec("dicTT%s=dictt%s1" % (i, i))
        exec("doc%s.render(dicTT%s)" % (i, i))
        exec("doc%s.save('试卷'+str(i)+'.docx')" % i)


root = tk.Tk()
root.geometry("830x480")
l1 = tk.Label(root, text="齐鲁工业大学印刷专业试卷出题系统：", bg="lightyellow", fg="red")
l1.grid(column=0, row=0)
v1, v2, v3, v4, v5, v6, v7 = tk.IntVar(), tk.IntVar(), tk.IntVar(), tk.IntVar(), tk.IntVar(), tk.IntVar(), tk.IntVar()

GUI("填空", v1, 1)
GUI("名词解释", v2, 2)
GUI("问答", v3, 3)
GUI("案例", v4, 4)
GUI("题型5", v5, 5)
GUI("题型6", v6, 6)
GUI("题型7", v7, 7)

EntryList1 = []
EntryList2 = []
for i in range(7):
    varE1 = 'EnT' + str(i)
    EntryList1.append(varE1)
    varE1 = tk.Entry(root)
    varE1.grid(row=i + 1, column=2, padx=5, pady=5, sticky='w')
    EntryList1[i] = varE1
for j in range(7):
    varE2 = 'EnT' + str(j + 4)
    EntryList2.append(varE2)
    varE2 = tk.Entry(root)
    varE2.grid(row=j + 1, column=4, padx=5, pady=5, sticky='w')
    EntryList2[j] = varE2

l_zong = tk.Label(root, text="试卷总分：", font=3)
l_zong.grid(column=1, row=8, ipadx=5, pady=40)
text = tk.Text(root, width=20, height=2)
text.grid(column=2, row=8)
cal_button = tk.Button(root, text='add!', command=cal_add, font=3).grid(column=3, row=8, ipadx=5)
button1 = tk.Button(root, text="随机抽题", command=twj, fg = "blue", font=3).grid(column=0, row=9, ipadx=5)
button3 = tk.Button(root, text="生成试卷", command=word, fg = "blue", font=3).grid(column=2, row=9, ipadx=5)
button4 = tk.Button(root, text="退出系统", command=sys.exit, fg = "blue", font=3).grid(column=4, row=9, ipadx=5)
root.mainloop()